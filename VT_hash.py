import requests
import openpyxl
import time

def hash_to_other_formats(api_key, hash_value):
    url = f'https://www.virustotal.com/vtapi/v2/file/report'
    params = {'apikey': api_key, 'resource': hash_value}
    response = requests.get(url, params=params)
    if response.status_code == 200:
        json_response = response.json()
        if json_response['response_code'] == 1:
            return json_response['md5'],  json_response['sha1'],  json_response['sha256']
        else:
            return "Hash not found in VirusTotal database"
    else:
        return "Error occurred while fetching data from VirusTotal"

# Replace 'YOUR_API_KEY' with your actual VirusTotal API key
api_key = '200957b097abd6c4c54b1b6d1ba25ddd27478be88861e8aedc16dc5c23addad4'

# Load Excel file
workbook = openpyxl.load_workbook('hashes.xlsx')
sheet = workbook.active
new_sheet = workbook.create_sheet("Converted_Hashes")
header = "md5","sha1","sha256"
new_sheet.append(header)

# Iterate through each row in the Excel file
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=4, values_only=True):
    hash_value = row[0]
    converted_hash = hash_to_other_formats(api_key, hash_value)
    new_sheet.append(converted_hash)
    print("Converting Hash.....!!!!")
    time.sleep(15)

# Save the modified Excel file
print("Successfuly Converted the hashes")
workbook.save('hashes_with_formats.xlsx')
